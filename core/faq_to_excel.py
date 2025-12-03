"""
Módulo para converter FAQ processado em planilha Excel.
"""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def parse_faq_text(faq_text):
    """
    Parse do texto FAQ estruturado em dicionários.
    
    Formato esperado:
    q: Pergunta
    sq: variação 1 | variação 2
    a: Resposta
    f:
    - Item 1
    - Item 2
    ...
    tags: palavra1, palavra2, palavra3
    categoria: uma das categorias válidas
    ---
    """
    faq_items = []
    current_item = {}
    current_f = []
    in_f_section = False
    
    lines = faq_text.split('\n')
    
    for line in lines:
        line = line.strip()
        
        if not line:
            continue
        
        # Detecta separador entre itens
        if line.startswith('---'):
            if current_item:
                if current_f:
                    current_item['f'] = '\n'.join(current_f)
                faq_items.append(current_item)
                current_item = {}
                current_f = []
                in_f_section = False
            continue
        
        # Detecta campos
        if line.startswith('q:'):
            in_f_section = False
            current_item['q'] = line[2:].strip()
        elif line.startswith('sq:'):
            in_f_section = False
            current_item['sq'] = line[3:].strip()
        elif line.startswith('a:'):
            in_f_section = False
            current_item['a'] = line[2:].strip()
        elif line.startswith('f:'):
            in_f_section = True
            current_f = []
        elif line.startswith('tags:'):
            in_f_section = False
            current_item['tags'] = line[5:].strip()
        elif line.startswith('categoria:'):
            in_f_section = False
            current_item['categoria'] = line[10:].strip()
        elif line.startswith('m:'):
            in_f_section = False
            # Ignora campo m (cadeia de pensamento) por enquanto
        elif line.startswith('t:'):
            in_f_section = False
            # Ignora campo t (trecho) por enquanto
        elif line.startswith('evidence_level:'):
            in_f_section = False
            # Ignora por enquanto
        elif line.startswith('confidence:'):
            in_f_section = False
            # Ignora por enquanto
        elif in_f_section and line.startswith('-'):
            current_f.append(line[1:].strip())
        elif in_f_section:
            # Continuação do item f
            if current_f:
                current_f[-1] += " " + line
    
    # Adiciona último item se existir
    if current_item:
        if current_f:
            current_item['f'] = '\n'.join(current_f)
        faq_items.append(current_item)
    
    return faq_items


def create_faq_excel(faq_text, output_path, source_name="Documento"):
    """
    Cria planilha Excel a partir do texto FAQ processado.
    
    Args:
        faq_text: Texto FAQ processado
        output_path: Caminho do arquivo Excel a ser criado
        source_name: Nome da fonte (para título)
    """
    # Parse do FAQ
    faq_items = parse_faq_text(faq_text)
    
    if not faq_items:
        raise ValueError("Nenhum item FAQ encontrado no texto. Verifique o formato.")
    
    # Cria workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "FAQ"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    
    # Título
    ws['A1'] = f"FAQ Completo - {source_name}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    # Cabeçalhos
    headers = ['ID', 'Pergunta Principal (q)', 'Variações de Pergunta (sq)', 'Resposta (a)', 'Framework/Detalhes (f)', 'Palavras-chave (tags)', 'Categoria']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados
    for idx, item in enumerate(faq_items, 1):
        row = idx + 3
        
        # ID
        ws.cell(row=row, column=1, value=idx)
        
        # Pergunta (q)
        ws.cell(row=row, column=2, value=item.get('q', ''))
        
        # Variações (sq)
        ws.cell(row=row, column=3, value=item.get('sq', ''))
        
        # Resposta (a)
        ws.cell(row=row, column=4, value=item.get('a', ''))
        
        # Framework (f)
        ws.cell(row=row, column=5, value=item.get('f', ''))
        
        # Palavras-chave (tags)
        tags = item.get('tags', '')
        ws.cell(row=row, column=6, value=tags)
        
        # Categoria (usa palavras-chave e conteúdo para determinar)
        category = item.get('categoria', '')
        if not category:
            # Se não houver categoria explícita, extrai das palavras-chave
            category = extract_category_from_keywords(
                tags=tags,
                f_text=item.get('f', ''),
                q_text=item.get('q', ''),
                a_text=item.get('a', '')
            )
        ws.cell(row=row, column=7, value=category)
        
        # Aplica formatação
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap_alignment
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 40  # Pergunta
    ws.column_dimensions['C'].width = 40  # Variações
    ws.column_dimensions['D'].width = 60  # Resposta
    ws.column_dimensions['E'].width = 60  # Framework
    ws.column_dimensions['F'].width = 30  # Palavras-chave
    ws.column_dimensions['G'].width = 20  # Categoria
    
    # Ajusta altura das linhas
    for row in range(4, len(faq_items) + 4):
        ws.row_dimensions[row].height = 60
    
    # Salva arquivo
    wb.save(output_path)
    return len(faq_items)


def extract_category_from_keywords(tags, f_text="", q_text="", a_text=""):
    """
    Extrai categoria baseada em palavras-chave (tags) e conteúdo.
    
    Args:
        tags: String com palavras-chave separadas por vírgula
        f_text: Texto do framework (opcional)
        q_text: Texto da pergunta (opcional)
        a_text: Texto da resposta (opcional)
    
    Returns:
        str: Categoria determinada
    """
    # Categorias válidas
    categories = [
        "Concept", "Principle", "Procedure", "Example", "Story", 
        "Quote", "Strategy", "Metric", "Warning", "Checklist", 
        "Decision", "Framework", "Exercise"
    ]
    
    # Combina todos os textos para análise
    all_text = " ".join([tags or "", f_text or "", q_text or "", a_text or ""]).lower()
    tags_lower = (tags or "").lower()
    
    # Primeiro, verifica se há categoria explícita no texto
    for cat in categories:
        if cat.lower() in all_text:
            return cat
    
    # Mapeamento de palavras-chave para categorias
    keyword_mapping = {
        "Procedure": [
            'procedimento', 'procedure', 'passo', 'step', 'como fazer', 'how to',
            'processo', 'process', 'método', 'method', 'técnica', 'technique',
            'instrução', 'instruction', 'guia', 'guide', 'tutorial', 'passo a passo'
        ],
        "Example": [
            'exemplo', 'example', 'caso', 'case', 'história', 'story', 'narrativa',
            'situação', 'situation', 'cenário', 'scenario', 'ilustração', 'illustration'
        ],
        "Principle": [
            'princípio', 'principio', 'principle', 'regra', 'rule', 'fundamento',
            'fundament', 'base', 'base', 'conceito base', 'core concept', 'lei', 'law'
        ],
        "Strategy": [
            'estratégia', 'strategy', 'tática', 'tactic', 'abordagem', 'approach',
            'plano', 'plan', 'metodologia', 'methodology', 'framework', 'estrutura'
        ],
        "Metric": [
            'métrica', 'metric', 'medida', 'measure', 'indicador', 'indicator',
            'kpi', 'performance', 'desempenho', 'resultado', 'result', 'avaliação'
        ],
        "Warning": [
            'aviso', 'warning', 'cuidado', 'care', 'atenção', 'attention', 'alerta',
            'alert', 'risco', 'risk', 'perigo', 'danger', 'evitar', 'avoid'
        ],
        "Checklist": [
            'checklist', 'lista', 'list', 'verificação', 'verification', 'itens',
            'items', 'pontos', 'points', 'checagem', 'check'
        ],
        "Concept": [
            'conceito', 'concept', 'definição', 'definition', 'entendimento',
            'understanding', 'noção', 'notion', 'ideia', 'idea'
        ],
        "Decision": [
            'decisão', 'decision', 'escolha', 'choice', 'seleção', 'selection',
            'opção', 'option', 'alternativa', 'alternative'
        ],
        "Framework": [
            'framework', 'estrutura', 'structure', 'modelo', 'model', 'arquitetura',
            'architecture', 'sistema', 'system'
        ],
        "Exercise": [
            'exercício', 'exercise', 'prática', 'practice', 'atividade', 'activity',
            'treino', 'training', 'simulação', 'simulation'
        ],
        "Quote": [
            'citação', 'quote', 'frase', 'phrase', 'dito', 'saying', 'máxima',
            'maxim', 'provérbio', 'proverb'
        ]
    }
    
    # Conta ocorrências de palavras-chave por categoria
    category_scores = {}
    for category, keywords in keyword_mapping.items():
        score = 0
        for keyword in keywords:
            # Pesos maiores para tags (palavras-chave)
            if keyword in tags_lower:
                score += 3
            if keyword in all_text:
                score += 1
        if score > 0:
            category_scores[category] = score
    
    # Retorna a categoria com maior score
    if category_scores:
        return max(category_scores, key=category_scores.get)
    
    # Fallback: análise por padrões no texto
    if any(word in all_text for word in ['passo', 'step', 'procedimento', 'procedure', 'como fazer']):
        return "Procedure"
    elif any(word in all_text for word in ['exemplo', 'example', 'caso', 'história']):
        return "Example"
    elif any(word in all_text for word in ['princípio', 'principio', 'principle', 'regra']):
        return "Principle"
    elif any(word in all_text for word in ['estratégia', 'strategy', 'tática', 'plano']):
        return "Strategy"
    elif any(word in all_text for word in ['métrica', 'metric', 'medida', 'indicador']):
        return "Metric"
    elif any(word in all_text for word in ['checklist', 'lista', 'verificação']):
        return "Checklist"
    elif any(word in all_text for word in ['aviso', 'warning', 'cuidado', 'atenção']):
        return "Warning"
    
    return "Concept"  # Default mais apropriado que "Geral"


def create_consolidated_faq_excel(faq_texts_list, source_names_list, output_path, output_language="pt"):
    """
    Cria planilha Excel consolidada com múltiplos FAQs.
    
    Args:
        faq_texts_list: Lista de textos FAQ processados
        source_names_list: Lista de nomes das fontes (mesma ordem)
        output_path: Caminho do arquivo Excel a ser criado
        output_language: Idioma (para título)
    
    Returns:
        int: Total de itens FAQ consolidados
    """
    if not faq_texts_list or not source_names_list:
        raise ValueError("Lista de FAQs ou nomes de fontes vazia")
    
    if len(faq_texts_list) != len(source_names_list):
        raise ValueError("Número de FAQs não corresponde ao número de nomes de fontes")
    
    # Parse de todos os FAQs
    all_faq_items = []
    
    for faq_idx, (faq_text, source_name) in enumerate(zip(faq_texts_list, source_names_list)):
        faq_items = parse_faq_text(faq_text)
        
        for item in faq_items:
            item['source'] = source_name  # Adiciona fonte ao item
            all_faq_items.append(item)
    
    if not all_faq_items:
        raise ValueError("Nenhum item FAQ encontrado nos textos fornecidos.")
    
    # Cria workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "FAQ Consolidado"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    source_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Título
    total_sources = len(set(source_names_list))
    ws['A1'] = f"FAQ Consolidado - {total_sources} Fonte(s)"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    
    # Informações das fontes
    ws['A2'] = "Fontes processadas:"
    ws['A2'].font = Font(bold=True)
    sources_text = " | ".join(set(source_names_list))
    ws.merge_cells('B2:H2')
    ws['B2'] = sources_text
    ws['B2'].alignment = wrap_alignment
    
    # Cabeçalhos
    headers = ['ID', 'Fonte', 'Pergunta Principal (q)', 'Variações de Pergunta (sq)', 'Resposta (a)', 'Framework/Detalhes (f)', 'Palavras-chave (tags)', 'Categoria']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados
    for idx, item in enumerate(all_faq_items, 1):
        row = idx + 4
        
        # ID
        ws.cell(row=row, column=1, value=idx)
        
        # Fonte (com cor de fundo alternada)
        source_cell = ws.cell(row=row, column=2, value=item.get('source', 'Desconhecida'))
        if idx % 2 == 0:
            source_cell.fill = source_fill
        
        # Pergunta (q)
        ws.cell(row=row, column=3, value=item.get('q', ''))
        
        # Variações (sq)
        ws.cell(row=row, column=4, value=item.get('sq', ''))
        
        # Resposta (a)
        ws.cell(row=row, column=5, value=item.get('a', ''))
        
        # Framework (f)
        ws.cell(row=row, column=6, value=item.get('f', ''))
        
        # Palavras-chave (tags)
        tags = item.get('tags', '')
        ws.cell(row=row, column=7, value=tags)
        
        # Categoria (usa palavras-chave e conteúdo para determinar)
        category = item.get('categoria', '')
        if not category:
            # Se não houver categoria explícita, extrai das palavras-chave
            category = extract_category_from_keywords(
                tags=tags,
                f_text=item.get('f', ''),
                q_text=item.get('q', ''),
                a_text=item.get('a', '')
            )
        ws.cell(row=row, column=8, value=category)
        
        # Aplica formatação
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap_alignment
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 25  # Fonte
    ws.column_dimensions['C'].width = 40  # Pergunta
    ws.column_dimensions['D'].width = 40  # Variações
    ws.column_dimensions['E'].width = 60  # Resposta
    ws.column_dimensions['F'].width = 60  # Framework
    ws.column_dimensions['G'].width = 30  # Palavras-chave
    ws.column_dimensions['H'].width = 20  # Categoria
    
    # Ajusta altura das linhas
    for row in range(5, len(all_faq_items) + 5):
        ws.row_dimensions[row].height = 60
    
    # Salva arquivo
    wb.save(output_path)
    return len(all_faq_items)

