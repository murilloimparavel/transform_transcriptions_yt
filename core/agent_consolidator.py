"""
Consolidador de bases de conhecimento para agentes de IA.
Combina múltiplos arquivos Agent Builder em uma estrutura unificada otimizada para RAG.
"""

import os
import sys
import json
import re
import shutil
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from dotenv import load_dotenv

load_dotenv()

# Fix encoding para Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Tenta importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl não instalado. Planilhas serão geradas em CSV.")


class AgentConsolidator:
    """
    Consolida múltiplos arquivos Agent Builder em uma estrutura unificada.
    """

    def __init__(self, project_name: str, output_language: str = "pt"):
        self.project_name = self._sanitize_name(project_name)
        self.output_language = output_language
        self.json_files: List[str] = []
        self.txt_files: List[str] = []
        self.consolidated_data = {
            "metadata": {
                "project_name": project_name,
                "created_at": datetime.now().isoformat(),
                "language": output_language,
                "sources": []
            },
            "qa_items": [],
            "facts": [],
            "procedures": [],
            "glossary": [],
            "examples": [],
            "agent_instructions": []
        }

    def _sanitize_name(self, name: str) -> str:
        """Sanitiza nome para uso em caminhos de arquivo."""
        # Remove caracteres especiais
        sanitized = re.sub(r'[<>:"/\\|?*]', '_', name)
        # Limita tamanho
        return sanitized[:50]

    def discover_files(self, source_dir: str, pattern: str = "agent_builder") -> Tuple[List[str], List[str]]:
        """
        Descobre arquivos JSON e TXT do Agent Builder em um diretório.

        Args:
            source_dir: Diretório onde buscar
            pattern: Padrão para identificar arquivos Agent Builder

        Returns:
            Tupla com listas de arquivos JSON e TXT encontrados
        """
        json_files = []
        txt_files = []

        if not os.path.exists(source_dir):
            print(f"⚠️  Diretório não encontrado: {source_dir}")
            return [], []

        for filename in os.listdir(source_dir):
            if pattern in filename:
                filepath = os.path.join(source_dir, filename)
                if filename.endswith('.json'):
                    json_files.append(filepath)
                elif filename.endswith('.txt'):
                    txt_files.append(filepath)

        self.json_files = sorted(json_files)
        self.txt_files = sorted(txt_files)

        print(f"📁 Encontrados: {len(self.json_files)} JSONs e {len(self.txt_files)} TXTs")
        return self.json_files, self.txt_files

    def extract_qa_from_content(self, content: str, source_name: str) -> List[Dict]:
        """
        Extrai pares de Q&A do conteúdo de um bloco.
        Usa múltiplos padrões para capturar diferentes formatos.
        """
        qa_items = []

        # Padrão 1: Formato qa_id estruturado
        qa_pattern = r'qa_id:\s*(\w+)\s*\n.*?pergunta:\s*(.+?)(?:\n|$).*?(?:variacoes:.*?\n)?.*?resposta:\s*(.+?)(?=\nqa_id:|\nqa_\d+:|$)'

        # Padrão 2: Formato pergunta/resposta simples
        simple_pattern = r'(?:pergunta|question|q):\s*(.+?)\n.*?(?:resposta|answer|a):\s*(.+?)(?=\n(?:pergunta|question|q):|\n##|\n---|\Z)'

        # Padrão 3: Formato numerado
        numbered_pattern = r'\*\*(?:Q|Pergunta)\s*\d*[:.]\*\*\s*(.+?)\n.*?\*\*(?:A|Resposta)[:.]\*\*\s*(.+?)(?=\n\*\*(?:Q|Pergunta)|\n##|\Z)'

        # Padrão 4: Formato com asteriscos
        asterisk_pattern = r'\*\s*\*\*(?:Pergunta|Question)[:.]\*\*\s*(.+?)\n.*?\*\s*\*\*(?:Resposta|Answer)[:.]\*\*\s*(.+?)(?=\n\*\s*\*\*(?:Pergunta|Question)|\n##|\Z)'

        found_items = set()  # Para evitar duplicatas

        for pattern in [qa_pattern, simple_pattern, numbered_pattern, asterisk_pattern]:
            matches = re.findall(pattern, content, re.DOTALL | re.IGNORECASE)
            for match in matches:
                if len(match) >= 2:
                    # Extrai pergunta e resposta
                    if len(match) == 3:
                        qa_id, question, answer = match
                    else:
                        question, answer = match[0], match[1]
                        qa_id = f"QA_{len(qa_items) + 1}"

                    # Limpa e valida
                    question = question.strip()[:500]
                    answer = answer.strip()[:2000]

                    # Evita duplicatas
                    key = (question[:100], answer[:100])
                    if key not in found_items and len(question) > 10 and len(answer) > 20:
                        found_items.add(key)
                        qa_items.append({
                            "id": qa_id,
                            "question": question,
                            "answer": answer,
                            "source": source_name,
                            "type": "qa"
                        })

        return qa_items

    def extract_facts_from_content(self, content: str, source_name: str) -> List[Dict]:
        """
        Extrai fatos e afirmações do conteúdo.
        """
        facts = []

        # Padrão para fatos estruturados
        fact_pattern = r'fato_id:\s*(\w+)\s*\n.*?afirmacao:\s*(.+?)(?:\n|$).*?tipo:\s*(\w+)'

        # Padrão para métricas
        metric_pattern = r'metrica:\s*(.+?)\n.*?valor:\s*(.+?)(?:\n|$)'

        # Padrão para bullets de fatos
        bullet_pattern = r'[-*]\s*\*\*(.+?):\*\*\s*(.+?)(?=\n[-*]|\n##|\Z)'

        found_items = set()

        # Extrai fatos estruturados
        for match in re.findall(fact_pattern, content, re.DOTALL | re.IGNORECASE):
            fact_id, statement, fact_type = match
            statement = statement.strip()[:1000]
            key = statement[:100]
            if key not in found_items and len(statement) > 20:
                found_items.add(key)
                facts.append({
                    "id": fact_id,
                    "statement": statement,
                    "type": fact_type.strip(),
                    "source": source_name
                })

        # Extrai métricas
        for match in re.findall(metric_pattern, content, re.DOTALL | re.IGNORECASE):
            metric_name, value = match
            statement = f"{metric_name.strip()}: {value.strip()}"
            key = statement[:100]
            if key not in found_items:
                found_items.add(key)
                facts.append({
                    "id": f"M_{len(facts) + 1}",
                    "statement": statement,
                    "type": "metrica",
                    "source": source_name
                })

        return facts

    def extract_procedures_from_content(self, content: str, source_name: str) -> List[Dict]:
        """
        Extrai procedimentos e processos do conteúdo.
        """
        procedures = []

        # Padrão para processos estruturados
        process_pattern = r'processo_id:\s*(\w+)\s*\n.*?nome:\s*(.+?)(?:\n|$).*?objetivo:\s*(.+?)(?:\n|$).*?passos:(.*?)(?=processo_id:|\Z)'

        # Padrão para passos numerados
        steps_pattern = r'(\d+)\.\s*(.+?)(?=\n\d+\.|\n##|\Z)'

        found_items = set()

        for match in re.findall(process_pattern, content, re.DOTALL | re.IGNORECASE):
            proc_id, name, objective, steps_text = match
            name = name.strip()[:200]
            objective = objective.strip()[:500]

            # Extrai passos
            steps = []
            for step_match in re.findall(steps_pattern, steps_text, re.DOTALL):
                step_num, step_text = step_match
                steps.append(f"{step_num}. {step_text.strip()[:300]}")

            key = name[:50]
            if key not in found_items:
                found_items.add(key)
                procedures.append({
                    "id": proc_id,
                    "name": name,
                    "objective": objective,
                    "steps": steps[:20],  # Limita a 20 passos
                    "source": source_name
                })

        return procedures

    def extract_glossary_from_content(self, content: str, source_name: str) -> List[Dict]:
        """
        Extrai termos do glossário do conteúdo.
        """
        glossary = []

        # Padrão para termos estruturados
        term_pattern = r'termo:\s*(.+?)(?:\n|$).*?definicao:\s*(.+?)(?=\ntermo:|\n##|\Z)'

        # Padrão alternativo
        alt_pattern = r'\*\*(.+?):\*\*\s*(.+?)(?=\n\*\*|\n##|\Z)'

        found_items = set()

        for pattern in [term_pattern, alt_pattern]:
            for match in re.findall(pattern, content, re.DOTALL | re.IGNORECASE):
                term, definition = match
                term = term.strip()[:100]
                definition = definition.strip()[:500]

                key = term.lower()
                if key not in found_items and len(term) > 2 and len(definition) > 10:
                    found_items.add(key)
                    glossary.append({
                        "term": term,
                        "definition": definition,
                        "source": source_name
                    })

        return glossary

    def extract_examples_from_content(self, content: str, source_name: str) -> List[Dict]:
        """
        Extrai exemplos e casos do conteúdo.
        """
        examples = []

        # Padrão para casos estruturados
        case_pattern = r'caso_id:\s*(\w+)\s*\n.*?titulo:\s*(.+?)(?:\n|$).*?(?:contexto|situacao).*?:\s*(.+?)(?=\ncaso_id:|\n##|\Z)'

        found_items = set()

        for match in re.findall(case_pattern, content, re.DOTALL | re.IGNORECASE):
            case_id, title, context = match
            title = title.strip()[:200]
            context = context.strip()[:1000]

            key = title[:50]
            if key not in found_items:
                found_items.add(key)
                examples.append({
                    "id": case_id,
                    "title": title,
                    "context": context,
                    "source": source_name
                })

        return examples

    def process_json_file(self, json_path: str) -> Dict:
        """
        Processa um arquivo JSON do Agent Builder e extrai dados estruturados.
        """
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        source_name = data.get('metadata', {}).get('source', os.path.basename(json_path))

        # Adiciona à lista de fontes
        self.consolidated_data['metadata']['sources'].append({
            "name": source_name,
            "file": os.path.basename(json_path),
            "generated_at": data.get('metadata', {}).get('generated_at', '')
        })

        # Extrai instruções do agente
        agent_instructions = data.get('agent_instructions', '')
        if agent_instructions:
            self.consolidated_data['agent_instructions'].append({
                "source": source_name,
                "instructions": agent_instructions
            })

        # Processa cada bloco de conhecimento
        knowledge_blocks = data.get('knowledge_blocks', {})

        for block_num, block_data in knowledge_blocks.items():
            block_name = block_data.get('name', '')
            block_content = block_data.get('content', '')

            # Bloco 1: Ontologia (Glossário)
            if 'ONTOLOGIA' in block_name.upper() or block_num == '1':
                glossary = self.extract_glossary_from_content(block_content, source_name)
                self.consolidated_data['glossary'].extend(glossary)

            # Bloco 2: Fatos
            if 'FACTUAL' in block_name.upper() or 'CONHECIMENTO' in block_name.upper() or block_num == '2':
                facts = self.extract_facts_from_content(block_content, source_name)
                self.consolidated_data['facts'].extend(facts)

            # Bloco 3: Procedimentos
            if 'PROCEDIMENTO' in block_name.upper() or 'INSTRUC' in block_name.upper() or block_num == '3':
                procedures = self.extract_procedures_from_content(block_content, source_name)
                self.consolidated_data['procedures'].extend(procedures)

            # Bloco 4: Exemplos
            if 'EXEMPLO' in block_name.upper() or 'CASO' in block_name.upper() or block_num == '4':
                examples = self.extract_examples_from_content(block_content, source_name)
                self.consolidated_data['examples'].extend(examples)

            # Bloco 5: Q&A
            if 'PERGUNTA' in block_name.upper() or 'Q&A' in block_name.upper() or 'RESPOSTA' in block_name.upper() or block_num == '5':
                qa_items = self.extract_qa_from_content(block_content, source_name)
                self.consolidated_data['qa_items'].extend(qa_items)

        return data

    def consolidate_all(self) -> Dict:
        """
        Processa todos os arquivos JSON descobertos e consolida os dados.
        """
        print("\n📊 Consolidando bases de conhecimento...")

        for json_path in self.json_files:
            try:
                print(f"  📄 Processando: {os.path.basename(json_path)}")
                self.process_json_file(json_path)
            except Exception as e:
                print(f"  ⚠️  Erro ao processar {json_path}: {e}")

        # Estatísticas
        stats = {
            "total_sources": len(self.consolidated_data['metadata']['sources']),
            "total_qa": len(self.consolidated_data['qa_items']),
            "total_facts": len(self.consolidated_data['facts']),
            "total_procedures": len(self.consolidated_data['procedures']),
            "total_glossary": len(self.consolidated_data['glossary']),
            "total_examples": len(self.consolidated_data['examples'])
        }

        self.consolidated_data['metadata']['stats'] = stats

        print(f"\n✅ Consolidação concluída:")
        print(f"   📚 Fontes: {stats['total_sources']}")
        print(f"   ❓ Q&As: {stats['total_qa']}")
        print(f"   📝 Fatos: {stats['total_facts']}")
        print(f"   📋 Procedimentos: {stats['total_procedures']}")
        print(f"   📖 Glossário: {stats['total_glossary']}")
        print(f"   💡 Exemplos: {stats['total_examples']}")

        return self.consolidated_data

    def generate_consolidated_system_prompt(self) -> str:
        """
        Gera um System Prompt consolidado a partir de todas as instruções.
        """
        if not self.consolidated_data['agent_instructions']:
            return "# System Prompt não disponível\n\nNenhuma instrução de agente foi encontrada."

        # Header
        prompt = f"""# SYSTEM PROMPT CONSOLIDADO
# Projeto: {self.project_name}
# Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}
# Fontes: {len(self.consolidated_data['metadata']['sources'])} documentos

---

## PERSONA DO AGENTE

Você é um especialista altamente qualificado no domínio coberto por este conhecimento consolidado.
Sua base de conhecimento foi construída a partir de {len(self.consolidated_data['metadata']['sources'])} fontes especializadas.

---

## ESTATÍSTICAS DA BASE DE CONHECIMENTO

- **Total de Q&As disponíveis**: {len(self.consolidated_data['qa_items'])}
- **Total de fatos documentados**: {len(self.consolidated_data['facts'])}
- **Total de procedimentos**: {len(self.consolidated_data['procedures'])}
- **Termos no glossário**: {len(self.consolidated_data['glossary'])}
- **Exemplos e casos**: {len(self.consolidated_data['examples'])}

---

## REGRAS DE RESPOSTA

1. Sempre baseie suas respostas no conhecimento documentado
2. Cite a fonte quando relevante
3. Se não souber, admita e sugira onde buscar
4. Use exemplos práticos sempre que possível
5. Mantenha consistência com o tom e terminologia das fontes

---

## INSTRUÇÕES ESPECÍFICAS POR FONTE

"""
        # Adiciona instruções de cada fonte
        for idx, instruction in enumerate(self.consolidated_data['agent_instructions'], 1):
            source = instruction.get('source', f'Fonte {idx}')
            content = instruction.get('instructions', '')

            # Limita tamanho de cada instrução
            if len(content) > 3000:
                content = content[:3000] + "\n\n[... conteúdo truncado ...]"

            prompt += f"""
### Fonte {idx}: {source}

{content}

---

"""

        # Footer com instruções finais
        prompt += f"""
## INSTRUÇÕES FINAIS

1. Priorize informações que aparecem em múltiplas fontes
2. Ao responder perguntas, verifique primeiro os Q&As consolidados
3. Para explicar conceitos, use o glossário como referência
4. Para guiar ações, consulte os procedimentos documentados
5. Ilustre respostas com os exemplos e casos disponíveis

---

**Este prompt foi gerado automaticamente pelo YouTube Transcription Processor - Agent Builder**
"""

        return prompt

    def create_mega_excel(self, output_path: str) -> str:
        """
        Cria uma mega planilha Excel com todas as informações consolidadas.
        """
        if not EXCEL_AVAILABLE:
            return self.create_mega_csv(output_path.replace('.xlsx', '.csv'))

        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def style_header(ws, headers):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def auto_width(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = min(len(str(cell.value)), 80)
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

        # Aba 1: Q&A (Principal para RAG)
        ws_qa = wb.active
        ws_qa.title = "Q&A"
        headers_qa = ["ID", "Pergunta", "Resposta", "Fonte", "Tags"]
        style_header(ws_qa, headers_qa)

        for row, item in enumerate(self.consolidated_data['qa_items'], 2):
            ws_qa.cell(row=row, column=1, value=item.get('id', f'QA_{row-1}'))
            ws_qa.cell(row=row, column=2, value=item.get('question', ''))
            ws_qa.cell(row=row, column=3, value=item.get('answer', ''))
            ws_qa.cell(row=row, column=4, value=item.get('source', ''))
            ws_qa.cell(row=row, column=5, value=item.get('tags', ''))
            for col in range(1, 6):
                ws_qa.cell(row=row, column=col).alignment = cell_alignment
                ws_qa.cell(row=row, column=col).border = thin_border

        auto_width(ws_qa)

        # Aba 2: Fatos
        ws_facts = wb.create_sheet("Fatos")
        headers_facts = ["ID", "Afirmação", "Tipo", "Fonte"]
        style_header(ws_facts, headers_facts)

        for row, item in enumerate(self.consolidated_data['facts'], 2):
            ws_facts.cell(row=row, column=1, value=item.get('id', f'F_{row-1}'))
            ws_facts.cell(row=row, column=2, value=item.get('statement', ''))
            ws_facts.cell(row=row, column=3, value=item.get('type', ''))
            ws_facts.cell(row=row, column=4, value=item.get('source', ''))
            for col in range(1, 5):
                ws_facts.cell(row=row, column=col).alignment = cell_alignment
                ws_facts.cell(row=row, column=col).border = thin_border

        auto_width(ws_facts)

        # Aba 3: Procedimentos
        ws_proc = wb.create_sheet("Procedimentos")
        headers_proc = ["ID", "Nome", "Objetivo", "Passos", "Fonte"]
        style_header(ws_proc, headers_proc)

        for row, item in enumerate(self.consolidated_data['procedures'], 2):
            steps_text = "\n".join(item.get('steps', []))
            ws_proc.cell(row=row, column=1, value=item.get('id', f'P_{row-1}'))
            ws_proc.cell(row=row, column=2, value=item.get('name', ''))
            ws_proc.cell(row=row, column=3, value=item.get('objective', ''))
            ws_proc.cell(row=row, column=4, value=steps_text)
            ws_proc.cell(row=row, column=5, value=item.get('source', ''))
            for col in range(1, 6):
                ws_proc.cell(row=row, column=col).alignment = cell_alignment
                ws_proc.cell(row=row, column=col).border = thin_border

        auto_width(ws_proc)

        # Aba 4: Glossário
        ws_gloss = wb.create_sheet("Glossário")
        headers_gloss = ["Termo", "Definição", "Fonte"]
        style_header(ws_gloss, headers_gloss)

        for row, item in enumerate(self.consolidated_data['glossary'], 2):
            ws_gloss.cell(row=row, column=1, value=item.get('term', ''))
            ws_gloss.cell(row=row, column=2, value=item.get('definition', ''))
            ws_gloss.cell(row=row, column=3, value=item.get('source', ''))
            for col in range(1, 4):
                ws_gloss.cell(row=row, column=col).alignment = cell_alignment
                ws_gloss.cell(row=row, column=col).border = thin_border

        auto_width(ws_gloss)

        # Aba 5: Exemplos
        ws_ex = wb.create_sheet("Exemplos")
        headers_ex = ["ID", "Título", "Contexto", "Fonte"]
        style_header(ws_ex, headers_ex)

        for row, item in enumerate(self.consolidated_data['examples'], 2):
            ws_ex.cell(row=row, column=1, value=item.get('id', f'E_{row-1}'))
            ws_ex.cell(row=row, column=2, value=item.get('title', ''))
            ws_ex.cell(row=row, column=3, value=item.get('context', ''))
            ws_ex.cell(row=row, column=4, value=item.get('source', ''))
            for col in range(1, 5):
                ws_ex.cell(row=row, column=col).alignment = cell_alignment
                ws_ex.cell(row=row, column=col).border = thin_border

        auto_width(ws_ex)

        # Aba 6: Metadados
        ws_meta = wb.create_sheet("Metadados")
        headers_meta = ["Propriedade", "Valor"]
        style_header(ws_meta, headers_meta)

        meta_rows = [
            ("Projeto", self.project_name),
            ("Data de Criação", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("Idioma", self.output_language),
            ("Total de Fontes", len(self.consolidated_data['metadata']['sources'])),
            ("Total de Q&As", len(self.consolidated_data['qa_items'])),
            ("Total de Fatos", len(self.consolidated_data['facts'])),
            ("Total de Procedimentos", len(self.consolidated_data['procedures'])),
            ("Total de Termos", len(self.consolidated_data['glossary'])),
            ("Total de Exemplos", len(self.consolidated_data['examples'])),
        ]

        for row, (prop, val) in enumerate(meta_rows, 2):
            ws_meta.cell(row=row, column=1, value=prop)
            ws_meta.cell(row=row, column=2, value=str(val))

        # Lista de fontes
        row_offset = len(meta_rows) + 3
        ws_meta.cell(row=row_offset, column=1, value="FONTES PROCESSADAS")
        ws_meta.cell(row=row_offset, column=1).font = Font(bold=True)

        for idx, source in enumerate(self.consolidated_data['metadata']['sources'], 1):
            ws_meta.cell(row=row_offset + idx, column=1, value=source.get('name', ''))
            ws_meta.cell(row=row_offset + idx, column=2, value=source.get('file', ''))

        auto_width(ws_meta)

        # Salva
        wb.save(output_path)
        return output_path

    def create_mega_csv(self, output_path: str) -> str:
        """
        Fallback: Cria CSV se openpyxl não estiver disponível.
        """
        import csv

        # Salva Q&As (principal)
        qa_path = output_path.replace('.csv', '_qa.csv')
        with open(qa_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Pergunta", "Resposta", "Fonte", "Tags"])
            for item in self.consolidated_data['qa_items']:
                writer.writerow([
                    item.get('id', ''),
                    item.get('question', ''),
                    item.get('answer', ''),
                    item.get('source', ''),
                    item.get('tags', '')
                ])

        print(f"  📄 CSV de Q&As salvo: {qa_path}")
        return qa_path

    def organize_output(self, base_output_dir: str) -> str:
        """
        Organiza todos os outputs em uma estrutura de pastas.

        Estrutura:
        base_output_dir/
        ├── mega_planilha/
        │   └── knowledge_base.xlsx
        ├── system_prompt/
        │   └── system_prompt.txt
        ├── json/
        │   └── [arquivos json originais]
        └── txt/
            └── [arquivos txt originais]
        """
        # Cria diretório base
        project_dir = os.path.join(base_output_dir, f"agent_consolidado_{self.project_name}")

        # Subdiretórios
        dirs = {
            'mega_planilha': os.path.join(project_dir, 'mega_planilha'),
            'system_prompt': os.path.join(project_dir, 'system_prompt'),
            'json': os.path.join(project_dir, 'json'),
            'txt': os.path.join(project_dir, 'txt')
        }

        # Cria diretórios
        for dir_path in dirs.values():
            os.makedirs(dir_path, exist_ok=True)

        print(f"\n📁 Organizando outputs em: {project_dir}")

        # 1. Gera e salva mega planilha
        excel_path = os.path.join(dirs['mega_planilha'], f"knowledge_base_{self.project_name}.xlsx")
        self.create_mega_excel(excel_path)
        print(f"  ✅ Mega planilha: {excel_path}")

        # 2. Gera e salva System Prompt
        system_prompt = self.generate_consolidated_system_prompt()
        prompt_path = os.path.join(dirs['system_prompt'], f"system_prompt_{self.project_name}.txt")
        with open(prompt_path, 'w', encoding='utf-8') as f:
            f.write(system_prompt)
        print(f"  ✅ System Prompt: {prompt_path}")

        # 3. Copia JSONs
        for json_file in self.json_files:
            dest = os.path.join(dirs['json'], os.path.basename(json_file))
            shutil.copy2(json_file, dest)
        print(f"  ✅ JSONs copiados: {len(self.json_files)} arquivos")

        # 4. Copia TXTs
        for txt_file in self.txt_files:
            dest = os.path.join(dirs['txt'], os.path.basename(txt_file))
            shutil.copy2(txt_file, dest)
        print(f"  ✅ TXTs copiados: {len(self.txt_files)} arquivos")

        # 5. Salva JSON consolidado
        consolidated_json_path = os.path.join(project_dir, f"consolidated_data_{self.project_name}.json")
        with open(consolidated_json_path, 'w', encoding='utf-8') as f:
            json.dump(self.consolidated_data, f, ensure_ascii=False, indent=2)
        print(f"  ✅ JSON consolidado: {consolidated_json_path}")

        return project_dir


def consolidate_agent_builder_outputs(
    source_dir: str = None,
    project_name: str = None,
    output_language: str = "pt"
) -> str:
    """
    Função principal para consolidar outputs do Agent Builder.

    Args:
        source_dir: Diretório onde estão os arquivos (default: data/processed)
        project_name: Nome do projeto (default: baseado no timestamp)
        output_language: Idioma de saída

    Returns:
        str: Caminho do diretório de saída
    """
    if source_dir is None:
        source_dir = os.path.join('data', 'processed')

    if project_name is None:
        project_name = f"projeto_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    print("\n" + "=" * 70)
    print("CONSOLIDADOR DE AGENT BUILDER")
    print("=" * 70)
    print(f"Diretorio fonte: {source_dir}")
    print(f"Nome do projeto: {project_name}")
    print(f"Idioma: {output_language}")
    print("=" * 70)

    # Inicializa consolidador
    consolidator = AgentConsolidator(project_name, output_language)

    # Descobre arquivos
    json_files, txt_files = consolidator.discover_files(source_dir, "agent_builder")

    if not json_files:
        print("❌ Nenhum arquivo Agent Builder encontrado!")
        return None

    # Consolida dados
    consolidator.consolidate_all()

    # Organiza outputs
    output_dir = consolidator.organize_output(source_dir)

    print("\n" + "=" * 70)
    print("✅ CONSOLIDAÇÃO CONCLUÍDA!")
    print("=" * 70)
    print(f"📁 Todos os arquivos organizados em: {output_dir}")
    print("=" * 70 + "\n")

    return output_dir
